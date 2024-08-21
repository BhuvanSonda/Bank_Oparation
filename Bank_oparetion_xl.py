import datetime as dt
import pandas as pd
from openpyxl import load_workbook

data = {"Names": [], "A/C No:": [], "Balance :": [], "Sign:": [], "Time:": []}
class Update:
    @staticmethod
    def update(name, account_no, sign, init_balance, time):
        data["Names"].append(name)
        data["A/C No:"].append(account_no)
        data["Balance :"].append(init_balance)
        data["Sign:"].append(sign)
        data["Time:"].append(time)
        df = pd.DataFrame(data)
        
        # Write DataFrame to Excel file
        try:
            with pd.ExcelWriter("data.xlsx", engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df.to_excel(writer, index=False, header=writer.sheets['Sheet1'].max_row == 0, startrow=writer.sheets['Sheet1'].max_row)
        except FileNotFoundError:
            with pd.ExcelWriter("data.xlsx", engine='openpyxl', mode='w') as writer:
                df.to_excel(writer, index=False)
try:
    read = pd.read_excel("data.xlsx")
except FileNotFoundError:
    read = pd.DataFrame(columns=["Names", "A/C No:", "Balance :", "Sign:", "Time:"])
class BankOperations:
    def __init__(self):
        try:
            self.read = pd.read_excel("data.xlsx")
        except FileNotFoundError:
            self.read = pd.DataFrame(columns=["Names", "A/C No:", "Balance :", "Sign:", "Time:"])

    def create(self, name, account_no, sign, init_balance):
        if account_no in self.read["A/C No:"].values:
            print("This account already exists.")
        else:
            time = dt.datetime.now()
            Update.update(name, account_no, sign, init_balance, time)
            print("Your account has been successfully created.")
            self.read = pd.read_excel("data.xlsx")

    def deposit(self, account_no, amount):
        if account_no in self.read["A/C No:"].values:
            index = self.read.index[self.read["A/C No:"] == account_no].tolist()[0]
            self.read.at[index, "Balance :"] += amount
            self.read.to_excel("data.xlsx", index=False)
            print("Your account has been successfully deposited.")
            print(f"Your account balance is {self.read.at[index, 'Balance :']}")
        else:
            print("Account number does not exist!")

    def withdraw(self, account_no, sign, amount):
        if account_no in self.read["A/C No:"].values:
            index = self.read.index[self.read["A/C No:"] == account_no].tolist()[0]
            if self.read.at[index, "Sign:"] == sign:
                if self.read.at[index, "Balance :"] >= amount:
                    self.read.at[index, "Balance :"] -= amount
                    self.read.to_excel("data.xlsx", index=False)
                    print("Your account has been successfully withdrawn.")
                    print(f"Your account balance is {self.read.at[index, 'Balance :']}")
                else:
                    print("Insufficient funds!")
            else:
                print("Mismatched signature!")
        else:
            print("Account number does not exist!")

    def check(self, account_no):
        if account_no in self.read["A/C No:"].values:
            index = self.read.index[self.read["A/C No:"] == account_no].tolist()[0]
            print(f"\n\nName : {self.read.at[index, 'Names']}")
            print(f"A/C No : {self.read.at[index, 'A/C No:']}")
            print(f"Available Balance : {self.read.at[index, 'Balance :']}")
        else:
            print("Account number does not exist!")
x = 1
bank_operations = BankOperations()

while x != 0:
    x = int(input("\nEnter\n0-->EXIT\n1-->Create new account\n2-->Deposit\n3-->Withdraw\n4-->Check balance\n"))

    if x == 0:
        break
    elif x == 1:
        name = input("Enter Name: ")
        account_no = input("Enter account number: ")
        init_balance = int(input("Enter amount to deposit: "))
        sign = input("Enter signature: ")
        bank_operations.create(name, account_no, sign, init_balance)
    elif x == 2:
        account_no = input("Enter account number: ")
        amount = int(input("Enter amount: "))
        bank_operations.deposit(account_no, amount)
    elif x == 3:
        account_no = input("Enter account number: ")
        amount = int(input("Enter amount: "))
        sign = input("Enter signature: ")
        bank_operations.withdraw(account_no, sign, amount)
    elif x == 4:
        account_no = input("Enter account number: ")
        bank_operations.check(account_no)
    else:
        print("Invalid input! Try again.")
