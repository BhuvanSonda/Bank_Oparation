import openpyxl 
from openpyxl.worksheet.table import Table, TableStyleInfo

workbook=openpyxl.load_workbook(r"C:\Users\Docketrun\Desktop\Bhuvan_python_program.py\18-09-2024\add_table.xlsx")
sheet=workbook.active


# Define the table (this assumes the table is already defined in the Excel file)
table = sheet.tables['Table1']  # Replace 'Table1' with your actual table name

# New data to append (ensure it matches the table structure)
new_data = [
    ['Banana',14250,2542,85632,4152],
    ['Grapes',1425,8542,7548,9653],
    ['Apple',1425,8542,7548,9653],
    ['Mango',1425,8542,7548,9653],
    ['Orange',1425,8542,7548,9653],
    ['Pineapple',1425,8542,7548,9653],
    ['Watermelon',1425,8542,7548,9653],
    ['Cherry',1425,8542,7548,9653],
]

# Find the last row of the existing table
last_row = table.ref.split(':')[1]
last_row_idx = int(last_row[1:])  # Get the row number

# Append new data to the table
for row in new_data:
    # Write data to the next empty row in the table
    for col_idx, value in enumerate(row, start=1):
        sheet.cell(row=last_row_idx + 1, column=col_idx+1, value=value)
    last_row_idx += 1  # Move to the next row

table.ref = f"{table.ref.split(':')[0]}:{sheet.cell(row=last_row_idx+1, column=len(row)+1).coordinate}"

# Reapply the table style if needed (optional)
style = TableStyleInfo(
    name="TableStyleMedium9", 
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=True
)
table.tableStyleInfo = style


workbook.save(r"C:\Users\Docketrun\Desktop\Bhuvan_python_program.py\18-09-2024\add_table.xlsx")