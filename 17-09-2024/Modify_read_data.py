import openpyxl as xl
import time

st_time=time.time()
workbook=xl.load_workbook(r"C:\Users\Docketrun\Desktop\Bhuvan_python_program.py\17-09-2024\Data_single_mul_mix.xlsx")
sheets=workbook.sheetnames
print(sheets)
sheet1=workbook[sheets[0]]

sheet1['B1']='Bhuvan'
sheet1['B2']=21
sheet1['A2']=20
sheet1['A3']='Honnavar'
sheet1['B3']='Sirsi'


sheet3=workbook[sheets[2]]
def find_cell_by_value(sheet, value):
    a=0
    # Iterate through all rows and cells to find the value
    for row in sheet.iter_rows():
        if a==1:
            break
        for cell in row:
            if cell.value == value:
                cell.value ='Nikhil'
                # print('Row = ',row)
                a=1
    else:
        print('Value not found!!')  # if the value is not found

# Example usage
value_to_find = 'Bhuvan'
cell = find_cell_by_value(sheet3, value_to_find)



workbook.save(r"C:\Users\Docketrun\Desktop\Bhuvan_python_program.py\17-09-2024\Data_single_mul_mix.xlsx")
print(f'time taken :{time.time()-st_time:.2f} seconds')