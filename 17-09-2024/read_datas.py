import openpyxl as xl
import os
import sys
import time
from openpyxl.drawing.image import Image
import numpy as np
from openpyxl.utils import get_column_letter
sys.path.insert(0, os.path.abspath(os.path.dirname("create_folder.py")))
from create_folder import Add_File

# Define the file name and path
file = 'Data_single_mul_mix1.xlsx'
file_path = Add_File(file)

st_time = time.time()
# Create a new workbook and add sheets
workbook = xl.Workbook()

# Sheet 1: Single Data
sheet1 = workbook.active
sheet1.title = 'single data'
sheet1['A1']='Santoshi'
sheet1['A2']=18
sheet1['A3']=12345
# Sheet 2: Multiple Data
sheet2 = workbook.create_sheet(title='multiple data')

# Data for 'multiple data' sheet
data_multiple = [
    ['Monitor', 100, 200], ['Pc', 120, 254],
    ['Mouse', 150, 300], ['Keyboard', 180, 350],
    ['Headset', 200, 400], ['Speaker', 220, 450],
    ['Printer', 250, 500], ['Scanner', 280, 550],
    ['Webcam', 310, 600], ['Tablet', 340, 650],
    ['Laptop', 370, 700], ['Desktop', 400, 750],
    ['Notebook', 430, 800], ['Camera', 460, 850],
    ['Microphone', 490, 900], ['Router', 520, 950],
    ['Switch', 550, 1000], ['Modem', 580, 1050]
]

# Add column headers and data to the 'multiple data' sheet
sheet2.append(['Items', 'Price', 'Sold_pieces'])
for row in data_multiple:
    sheet2.append(row)

# Sheet 3: Mixed Data
sheet3 = workbook.create_sheet(title='mixed data')

# Data for 'mixed data' sheet
data_mixed = [
    ['Bhuvan', 500, "PUC_II"],
    ["Radha", 180, "SSLC"],
    ['Girija', 250, "B.Com"],
    ['Rahul', 220, "B.Sc"],
    ['Ramesh', 280, "MBA"],
    ['Sushan', 320, 'MBBS'],
    ['Suresh', 350, 'MCA'],
    ['Suraj', 350, 'MCA'],
    ['Suresh', 420, 'MBA'],
    ['Ragu', 240, 'SSLC'],
    ['Rajesh', 260, 'B.Com'],
    ['Raj', 230, 'B.Com'],
    ['Santhoshi', 230, 'PUC_II']
]

# Add data to the 'mixed data' sheet
for row in data_mixed:
    sheet3.append(row)

# Determine the last column by checking the first row
last_column = sheet3.max_column
last_column_letter = get_column_letter(last_column)

# Identify the next column letter
next_column_letter = get_column_letter(last_column + 1)

# Define image dimensions  
def resize_image_to_fit_cell(image):
    scale = 0.5#min(max_width / image.width, max_height / image.height)
    image.width = int(image.width * scale)
    image.height = int(image.height * scale)
    width = int(image.width * scale)
    height =int(image.height * scale)
    return width, height

last_row = sheet3.max_row

# Add images to the next column
for row_number in range(2, last_row + 1, 2):  # Start from row 2 and step by 2
    immg = Image('openCV/cat.jpg') 

    # Set the width and height  of the images
    max_width = 15  
    max_height = 20 

#get the width and the height
    (max_width, max_height)=resize_image_to_fit_cell(immg)
    
    cell_position = f"{next_column_letter}{row_number}"
    immg.anchor = cell_position
    sheet3.add_image(immg)
    
    # Adjust row height and column width to fit images
    sheet3.column_dimensions[next_column_letter].width = max_width /2.5 
    sheet3.row_dimensions[row_number].height = max_height*1.5  


workbook.save(file_path)

print(f"Time taken: {time.time() - st_time:.2f} seconds")
