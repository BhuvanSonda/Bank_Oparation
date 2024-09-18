# # Import package and it's modules 
# from tkinter import *


# # text_update function 
# def text_updation(language): 
# 	text.delete(0, END) 
# 	text.insert(0, language) 

# # create root window 
# root = Tk() 

# # root window title and dimension 
# root.title("GeekForGeeks") 

# # Set geometry (widthxheight) 
# root.geometry('400x400') 

# # Entry Box 
# text = Entry(root, width=30, bg='White') 
# text.pack(pady=10) 

# # create buttons 
# button_dict = {} 
# words = ["Python", "Java", "R", "JavaScript"] 
# for lang in words: 
	
# 	# pass each button's text to a function 
# 	def action(x = lang): 
# 		return text_updation(x) 
		
# 	# create the buttons 
# 	button_dict[lang] = Button(root, text = lang, 
# 							command = action) 
# 	button_dict[lang].pack(pady=10) 

# # Execute Tkinter 
# root.mainloop()
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import time

# Create a workbook and add a sheet
workbook = Workbook()
file_path = 'your_workbook.xlsx'
sheet3 = workbook.create_sheet(title='mixed data')

# Add an image
immg = Image('famer.jpg')

# Data for 'mixed data' sheet
data_mixed = [
    ['Bhuvan', 500, "PUC_II"],
    ["Radha", 180, "SSLC"],
    ['Girija', 250, "B.Com"],
    ['Rahul', 220, "B.Sc"],
    ['Ramesh', 280, "MBA"],
    ['Sushan', 320, 'MBBS'],
    ['Suresh', 350, 'MCA'],
    ['Suraj', 350, 'MCA'],
    ['Suresh', 420, 'MBA'],
    ['Ragu', 240, 'SSLC'],
    ['Rajesh', 260, 'B.Com'],
    ['Raj', 230, 'B.Com'],
    ['Santhoshi', 230, 'PUC_II']
]

# Add data to the 'mixed data' sheet
for row in data_mixed:
    sheet3.append(row)

images = ['famer.jpg', 'famer.jpg', 'famer.jpg']

last_column = sheet3.max_column
last_column_letter = sheet3.cell(row=1, column=last_column).column_letter
for row_number, image_path in enumerate(images, start=2):  # start=2 to skip the header row
    immg = Image(image_path)
    cell_position = f"{last_column_letter}{row_number}"
    immg.anchor = cell_position
    sheet3.add_image(immg)


# Record start time and save the workbook
st_time = time.time()
workbook.save(file_path)
print(f"Time taken: {time.time() - st_time:.2f} seconds")

# Optional: Verify data
wb_check = load_workbook(file_path)
sheet3_check = wb_check['mixed data']
for row in sheet3_check.iter_rows(values_only=True):
    print(row)
