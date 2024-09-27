import openpyxl as xl
import sys
import os
import time
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
sys.path.insert(0, os.path.abspath(os.path.dirname('create_folder.py')))

from create_folder import Add_File 

file = "Write_image_data.xlsx"
file_path = Add_File(file)#create and  add file to the folder 

start_time=time.time()

workbook=xl.Workbook()#create workbook


sheet=workbook.active
image=['Ba.jpg','Bank_photo.jpg']*3


# Data for 'mixed data' sheet
data= [
    ['Bhuvan', 500, "PUC_II"],
    ["Radha", 180, "SSLC"],
    ['Girija', 250, "B.Com"],
    ['Rahul', 220, "B.Sc"],
    ['Ramesh', 280, "MBA"],
    ['Sushan', 320, 'MBBS'],
    
]

for row in data:
    sheet.append(row)

# Determine the last column by checking the first row
last_column = sheet.max_column
last_column_letter = get_column_letter(last_column)

# Identify the next column letter
next_column_letter = get_column_letter(last_column + 1)

def image_to_fit_cell(image):
    scale = 0.5
    image.width = int(image.width * scale)
    image.height = int(image.height * scale)
    width = int(image.width * scale)
    height =int(image.height * scale)
    return width, height
# Number of rows to add images to
last_row = sheet.max_row

# Add images to the next column
for row,img in enumerate(image):  # Start from row 2 and step by 2
    immg = Image(img)  # Path to the image

    # Resize image to fit within the cell
    (max_width, max_height)=image_to_fit_cell(immg)
    
    cell_position = f"{next_column_letter}{row}"
    immg.anchor = cell_position
    sheet.add_image(immg)
    
    # Adjust row height and column width to fit images
    sheet.column_dimensions[next_column_letter].width = max_width /2.5 
    sheet.row_dimensions[row].height = max_height*1.5   

# Record start time and save the workbook

workbook.save(file_path)

print(f"Time taken: {time.time() - start_time:.2f} seconds")
