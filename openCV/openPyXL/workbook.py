import openpyxl as xl
import pandas as pd

wb=xl.Workbook()
ws=wb.active

ws["A1"]='Name'
ws["B1"]='Age'
ws["C1"]='Gender'
ws['A2']='Bhuvan'
ws['B2']=25
ws['C2']='Male'

ws['A3']='Santoshi'
ws['B3']=30
ws['C3']='Female'

wb.save('save.xlsx')

