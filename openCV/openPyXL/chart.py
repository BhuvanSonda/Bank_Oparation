import openpyxl as xl

# Load the workbook and select the active sheet
wb = xl.load_workbook('save.xlsx')
sheet = wb.active

# Merge cells D1:E2
sheet.merge_cells('D1:E2')

# Set cell values with formulas
cell = sheet.cell(7, 2,'=SUM(B1:B6)')


cell2 = sheet.cell(7, 3)
cell2.value = '=SUMIF(C1:C6, "male", B1:B6)'  #   formula if you want a conditional sum

# Create a bar chart
chart = xl.chart.BarChart()
values = xl.chart.Reference(sheet, min_row=1, 
                            max_row=7, min_col=2,
                            max_col=3)      
print(values)
# Add data and title from the data
chart.add_data(values, titles_from_data=True)

# Add the chart to the sheet at position E2
sheet.add_chart(chart, "E20")

# Save the workbook with the changes
wb.save('save.xlsx')
