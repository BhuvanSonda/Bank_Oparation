import openpyxl as xl
#Load the workbook
wb = xl.load_workbook("save.xlsx",)

# Select the active worksheet
ws = wb.active
print(ws)

ws['A4']='Nagarathna'
ws['B4']=45
ws['C4']='Female'

# Modify cell values
ws.cell(row=2, column=1).value = 'Bhuvan'
ws.cell(row=2, column=2).value = 28

datas=(('sham',28,'male'),('Radha',23,'Female'))
# Insert data into the worksheet
for data in datas:
    ws.append(data)

# Save the workbook
wb.save("save.xlsx")