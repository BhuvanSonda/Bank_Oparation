import openpyxl as xl

wb=xl.load_workbook('save.xlsx')
ws=wb.active
print("\n************ Row iteration **********\n")
for row in ws.iter_rows(1,6,0,3):
    for cell in row:
        print(cell.value,"  ",end="")
    print("\n")
    
print("\n********** Column iteration **********\n")

for row in ws.iter_cols(0,3,1,6):
    # ws.append(row)
    for cell in row:
        print(cell.value,"  ",end="")
    print("\n")
    wb.save('save.xlsx')