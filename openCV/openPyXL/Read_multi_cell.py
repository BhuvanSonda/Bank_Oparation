import openpyxl as xl

wb=xl.load_workbook('save.xlsx')
ws=wb.active

cell_range=ws['A1':'B6'] #cell range 

for i1,i2, in cell_range:
    print("{0:6}{1:6}".format(i1.value,i2.value))

cell_range=ws['A1':'C6'] #cell range 

for i1,i2, in cell_range:
    print("{0:6}{1:6}{2:6}".format(i1.value,i2.value))