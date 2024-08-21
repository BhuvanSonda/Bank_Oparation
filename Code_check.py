import pandas as pd
import re
import datetime as dt
from openpyxl import load_workbook

class Update :
    def update(name,account_no,sign,init_balance,mobile,adhar,time):
        new_entry={'Name':[name],'A/C No':[account_no],
                   'Sign':[sign],
                   'Balance':[init_balance],'Mobile_No':[mobile],
                   'Adhar_No':[adhar],'Time':[time]}
        df=pd.DataFrame(new_entry)
        try:
           with pd.ExcelWriter('Bank.xlsx',engine='openpyxl',mode='a+',if_sheet_exists='overlay')as writer:
               df.to_excel(writer,sheet_name='Sheet1',index=False,header=None)
               try:
                    book = load_workbook('Bank.xlsx')
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    
                    # Get the last row in the existing Excel sheet
                    startrow = writer.sheets['Sheet1'].max_row

                    df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=startrow)
               except:
                    with pd.ExcelWriter('Bank.xlsx',engine='openpyxl',mode='w+')as writer:
                        df.to_excel(writer,sheet_name='Sheet1',index=False)
                        
        except Exception as e:
            print(f"An error occurred: {e}")
try:
    read=pd.read_excel('Bank.xlsx')
except FileNotFoundError:
    read=pd.DataFrame(columns=['Name','A/C No','Sign','Mobile_No','Adhar_No','Time'])

class BankOperation:
    try:
        read=pd.read_excel('Bank.xlsx')
    except FileNotFoundError:
        read=pd.DataFrame(columns=['Name','A/C No','Sign','Mobile_No','Adhar_No','Time'])
    
 

class Validation:
    def Mobile():
        M=re.compile("^[6-9]{1}[0-9]{9}$")
        mobile = input("Enter Mobile No:")
        if M.match(mobile):
            return mobile
        else:
            print("Invalid Mobile No")
            return Validation.Mobile()
    def Adhar():
        A = re.compile("^[2-9]{1}[0-9]{11}$")
        adhar = input("Enter Adhar No:")
        if A.match(adhar):
            return adhar
        else:
            print("Invalid Adhar No")
            return Validation.Adhar()
        

class BankOperations:
    def __init__(self):
        self.read = pd.read_excel("Bank.xlsx") if pd.io.common.file_exists("Bank.xlsx") else pd.DataFrame(columns=["Names", "A/C No", "Balance", "Sign","Mobile_No",'Adhar_No',"Time"])
    
    def ac_generator(self):
        self.read = pd.read_excel("Bank.xlsx") if pd.io.common.file_exists("Bank.xlsx") else pd.DataFrame(columns=["Names", "A/C No", "Balance", "Sign","Mobile_No",'Adhar_No',"Time"])
    
        if self.read["A/C No"].empty:
            return 100
        else:
            ac=0
            for i in self.read["A/C No"].values:
                ac=i
            return (ac+1)
            

    def create(self):
        name = input("Enter Name: ")
        account_no=self.ac_generator()
        init_balance = int(input("Enter amount to deposit: "))
        sign = input("Enter signature: ")
        mobile_no = Validation.Mobile()
        adhar_no = Validation.Adhar()

        if account_no in self.read["A/C No"].values:
            print("--"*15)
            print("This account already exists.")
           
        else:
            time = dt.datetime.now()
            Update.update(name, account_no, sign, init_balance,mobile_no,adhar_no, time)
            print("--"*15)
            print(f"Account created for--> {name} with A/C No: --> {account_no}")
            print("Your account has been successfully created.")
            print("--"*15)

           

    def deposit(self):
        account_no = int(input("Enter account number: "))
        count=0
        while True:
            if account_no in self.read["A/C No"].values:
                amount = int(input("Enter amount: "))
                index = self.read.index[self.read["A/C No"] == account_no].tolist()[0]
                self.read.at[index, "Balance"] += amount
                self.read.to_csv("Bank.xlsx", index=False)
                print("--"*15)
                print("Your account has been successfully deposited.")
                print(f"Your account balance is {self.read.at[index, 'Balance']}")
                break
            else:
                account_no = int(input("Enter Valid account number: "))
                count+=1
                if count==2:
                    print("--"*15)
                    print("Account number does not exist!")
                    choice=input("Please create your acount\npress\nY-->for create\nN-->for Not creat\n").lower()
                    if choice=='y':
                        self.create()
                        break
                    else:
                        print("--"*15)
                        print("Thank you for using our services")
                        break
                else:
                    continue

    
    def withdraw(self):
        account_no = int(input("Enter account number: "))
        count=0
        while True:
            if account_no in self.read["A/C No"].values:
                index = self.read.index[self.read["A/C No"] == account_no].tolist()[0]

                sign = input("Enter signature: ")
                count=0
                while self.read.at[index, "Sign"] != sign:
                    count+=1
                    if count==3:
                        print("--"*15)
                        print("Your account has been locked due to 3 wrong attempts")
                        break
                    print("--" * 15)
                    print("Mismatched signature!")
                    sign = input("Enter your correct signature: ")
                if count==3:
                    break

                if self.read.at[index, "Sign"] == sign:
                    amount = int(input("Enter amount: "))
                    if self.read.at[index, "Balance "] >= amount:
                        self.read.at[index, "Balance"] -= amount
                        self.read.to_csv("Bank.xlsx", index=False)
                        print("--"*15)
                        print("Your account has been successfully withdrawn.")
                        print(f"Your account balance is {self.read.at[index, 'Balance']}")
                        break
                    else:
                        print("--"*15)
                        print("Insufficient funds!")
                        count=0
                        while True:
                            count+=1
                            if count==3:
                                print("--"*15)
                                print("Your account has been locked due to 3 wrong attempts")
                                break
                            try:
                                amount = int(input("Enter valid amount: "))
                                if amount > 0 and self.read.at[index, "Balance"] >= amount:
                                    self.read.at[index, "Balance"] -= amount
                                    self.read.to_csv("Bank.xlsx", index=False)
                                    print("--" * 15)
                                    print("Your account has been successfully withdrawn.")
                                    print(f"Your account balance is {self.read.at[index, 'Balance']}")
                                    break
                                else:
                                    print("Insufficient funds or invalid amount! Please try again.")
                            except ValueError:
                                print("Invalid input. Please enter a valid Amount.")
                        if count==3:
                            break
            else:
                account_no = int(input("Enter Valid account number: "))
                count+=1
                if count==2:
                    print("--"*15)
                    print("Account number does not exist!")
                    choice=input("Please create your acount\npress\nY -->for create\nN-->for Not creat\n").lower()
                    if choice=='y':
                        self.create()
                        break
                    else:
                        print("--"*15)
                        print("Thank you for using our services")
                        break
                else:
                    continue        
        


    def check(self):
        account_no = int(input("Enter account number: "))
        count=0
        while True:
            if account_no in self.read["A/C No"].values:
                index = self.read.index[self.read["A/C No"] == account_no].tolist()[0]
                print("--"*15)
                print(f"\n\nName : {self.read.at[index, 'Name']}")
                print(f"A/C No : {self.read.at[index, 'A/C No']}")
                print(f"Available Balance : {self.read.at[index, 'Balance']}")
                print("--"*15)
                break
            
            else:
                account_no = int(input("Enter Valid account number: "))
                count+=1
                if count==2:
                    print("--"*15)
                    print("Account number does not exist!")
                    choice=input("Please create your acount")
                    print("--"*15)
                    print("\npress\nY -->for create\nN-->for Not creat\n").lower()
                    if choice=='y':
                        self.create()
                        break
                    else:
                        print("--"*15)
                        print("Thank you for using our services")
                        break
                else:
                    continue        
print("--"*10)

x = 1
bank_operations = BankOperations()

while x != 0:
    x = int(input("\nEnter\n0-->EXIT\n1-->Create new account\n2-->Deposit\n3-->Withdraw\n4-->Check balance\n"))
    print("--"*15)

    if x == 0:
        break
    
    elif x == 1:
        bank_operations.create()
    elif x == 2:
        bank_operations.deposit()
    elif x == 3:
        bank_operations.withdraw()
    elif x == 4:
        bank_operations.check()
    else:
        print("Invalid input! Try again.")
